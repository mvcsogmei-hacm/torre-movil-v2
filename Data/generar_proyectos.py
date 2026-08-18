# -*- coding: utf-8 -*-
"""
Genera Data/proyectos.js a partir del Excel consolidado.

Uso:  py generar_proyectos.py  (desde la carpeta Data)

Reglas de pertenencia a carteras (definidas por el usuario, 16-ago-2026):
  - paralizadas:    ESTADO SSP = PARALIZADA o PARALIZADO
  - transferencias: MONTO TOTAL TRANSFERIDO 2026 > 0
  - preset:         ETAPA DE EVALUACIÓN distinta de vacío y distinta de FINANCIADO
  - obras:          MODALIDAD DE FINANCIAMIENTO INTERNA = DIRECTA
"""
import json
import datetime
import openpyxl

EXCEL = "31.07.2026 - Matriz Única de Monitoreo - Consolidado.xlsx"
SALIDA = "proyectos.js"
HOJA = "CONSOLIDADO"


def limpio(v):
    """None, '', '-' -> None; resto como texto sin espacios sobrantes."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s not in ("", "-") else None


def num(v):
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0
    return int(f) if f == int(f) else round(f, 2)


def pct(v):
    """El Excel guarda el avance como fracción (0.8 = 80%)."""
    try:
        return round(float(v) * 100, 1)
    except (TypeError, ValueError):
        return 0


def fecha(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%d/%m/%Y")
    return limpio(v)


def main():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[HOJA]

    proyectos = []
    vistos = {}
    duplicados = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        cui = str(r[0]).strip()
        if cui in vistos:
            duplicados.append(cui)
            continue
        vistos[cui] = True

        p = {
            "cui": cui,
            "nombre": limpio(r[2]) or "",
            "programa": limpio(r[1]) or "",
            "uei": limpio(r[3]) or "",
            "dep": limpio(r[4]) or "",
            "prov": limpio(r[5]) or "",
            "dist": limpio(r[6]) or "",
            "modalidad": limpio(r[7]),
            "tipo": limpio(r[8]),
            "pobl": num(r[9]),
            "cxAgua": num(r[10]),
            "cxAlc": num(r[11]),
            "ssi": limpio(r[12]),
            "monto": num(r[13]),
            "devAc": num(r[14]),
            "pim": num(r[15]),
            "dev": num(r[16]),
            "fisico": pct(r[24]),
            "estadoET": limpio(r[19]),
            "procSel": limpio(r[20]),
            "ssp": limpio(r[21]),
            "subSsp": limpio(r[22]),
            "fTerm": fecha(r[23]),
        }

        carteras = {}

        etapa = limpio(r[17])
        if etapa is not None and etapa.upper() != "FINANCIADO":
            carteras["preset"] = {"etapa": etapa, "estado": limpio(r[18])}

        if (p["modalidad"] or "").upper() == "DIRECTA":
            carteras["obras"] = {"avance": p["fisico"]}

        transferido = num(r[27])
        if transferido > 0:
            carteras["transferencias"] = {
                "transferido": transferido,
                "ejecutado": num(r[28]),
            }

        if (p["ssp"] or "").upper() in ("PARALIZADA", "PARALIZADO"):
            carteras["paralizadas"] = {
                "avance": p["fisico"],
                "hito": limpio(r[25]),
                "fecha": fecha(r[26]),
            }

        if carteras:
            p["carteras"] = carteras
        proyectos.append(p)

    if duplicados:
        print("ADVERTENCIA: %d CUI duplicados omitidos: %s" % (len(duplicados), duplicados[:10]))

    # Indicadores de títulos de propiedad (hoja INDICADORES TITULOS)
    indicadores = {}
    if "INDICADORES TITULOS" in wb.sheetnames:
        titulos = {}
        for r in wb["INDICADORES TITULOS"].iter_rows(values_only=True):
            etiqueta = str(r[1] or "").strip().upper()
            if etiqueta.startswith("META"):
                titulos["meta"] = num(r[2])
            elif etiqueta.startswith("TITULOS ENTREGADOS") or etiqueta.startswith("TÍTULOS ENTREGADOS"):
                titulos["entregados"] = num(r[2])
        if titulos:
            indicadores["titulos"] = titulos

    def num_celda(v):
        """Número entero desde celda que puede venir como texto con espacio
        duro (\\xa0) o comas: '\\xa09028' -> 9028."""
        if v is None:
            return None
        if isinstance(v, str):
            v = v.replace("\xa0", "").replace(",", "").strip()
            if not v:
                return None
        try:
            return num(v)
        except (TypeError, ValueError):
            return None

    def pct_celda(v):
        """Normaliza el % de las hojas de títulos: '52,1%' -> 52.1;
        fracción (0.56) -> 56; número >1 (22.9) ya es porcentaje."""
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip().replace("%", "").replace(",", ".")
            try:
                v = float(s)
            except ValueError:
                return None
            return round(v, 1)
        v = float(v)
        return round(v * 100, 1) if v <= 1 else round(v, 1)

    # Macro regiones (hoja MACROREGION TITULOS): %, ejecución y meta
    if "MACROREGION TITULOS" in wb.sheetnames and "titulos" in indicadores:
        macro = []
        for r in wb["MACROREGION TITULOS"].iter_rows(values_only=True):
            nombre, val = limpio(r[1]), pct_celda(r[2])
            if nombre and val is not None:
                fila = {"nombre": nombre, "pct": val}
                ejec, meta = num_celda(r[3]), num_celda(r[4])
                if ejec is not None:
                    fila["ejec"] = ejec
                if meta is not None:
                    fila["meta"] = meta
                macro.append(fila)
        indicadores["titulos"]["macro"] = macro

    # Bonos (hojas INDICADORES BONOS, MODALIDAS BONOS, POR REGION BONOS)
    if "INDICADORES BONOS" in wb.sheetnames:
        bonos = {}
        for r in wb["INDICADORES BONOS"].iter_rows(values_only=True):
            etiqueta = str(r[1] or "").strip().upper()
            if etiqueta.startswith("META"):
                bonos["meta"] = num(r[2])
            elif etiqueta.startswith("DESEMBOLSADOS"):
                bonos["desembolsados"] = num(r[2])
        if bonos:
            indicadores["bonos"] = bonos
    def filas_con_detalle(hoja):
        """Filas nombre + % + desembolso/meta (hojas de bonos)."""
        filas = []
        for r in wb[hoja].iter_rows(values_only=True):
            nombre, val = limpio(r[1]), pct_celda(r[2])
            if not nombre or val is None:
                continue
            fila = {"nombre": nombre, "pct": val}
            des, meta = num_celda(r[3]), num_celda(r[4])
            if des is not None:
                fila["desembolso"] = des
            if meta is not None:
                fila["meta"] = meta
            filas.append(fila)
        return filas

    if "MODALIDAS BONOS" in wb.sheetnames and "bonos" in indicadores:
        indicadores["bonos"]["modalidades"] = filas_con_detalle("MODALIDAS BONOS")
    if "POR REGION BONOS" in wb.sheetnames and "bonos" in indicadores:
        indicadores["bonos"]["regiones"] = filas_con_detalle("POR REGION BONOS")

    # Indicador principal de Inicio (hoja INDICADOR PRINCIPAL): fila SECTOR → %
    if "INDICADOR PRINCIPAL" in wb.sheetnames:
        for r in wb["INDICADOR PRINCIPAL"].iter_rows(values_only=True):
            nombre, val = limpio(r[1]), pct_celda(r[2])
            if nombre and val is not None:
                indicadores["principal"] = {"nombre": nombre, "pct": val}
                break

    # Pliegos (hoja PLIEGOS): tarjetas de Inicio — entidad, PIM, devengado, % ejecución
    if "PLIEGOS" in wb.sheetnames:
        pliegos = []
        for r in wb["PLIEGOS"].iter_rows(values_only=True):
            nombre, pim, dev = limpio(r[2]), num(r[3]), num(r[4])
            val = pct_celda(r[5])
            if val is None and pim:          # la celda es fórmula: calcular
                val = round((dev or 0) / pim * 100, 1)
            if nombre and val is not None and nombre.upper() != "ENTIDAD":
                pliegos.append({"nombre": nombre, "pim": pim, "devengado": dev, "pct": val})
        if pliegos:
            indicadores["pliegos"] = pliegos

    # Detalle de Pliego MVCS (hoja MVCS): fila PLIEGO = resumen;
    # las demás filas = un card por fila en el detalle
    if "MVCS" in wb.sheetnames:
        filas_pliego = []
        for r in wb["MVCS"].iter_rows(values_only=True):
            nombre = limpio(r[1])
            todo = pct_celda(r[2])
            if not nombre or todo is None:
                continue
            filas_pliego.append({
                "nombre": nombre,
                "todo": todo,
                "actividades": pct_celda(r[3]),
                "proyectos": pct_celda(r[4]),
            })
        if filas_pliego:
            indicadores["pliego"] = filas_pliego

    # Entidades adscritas (hoja ADSCRITAS): nombre, % y monto
    if "ADSCRITAS" in wb.sheetnames:
        adscritas = []
        for r in wb["ADSCRITAS"].iter_rows(values_only=True):
            nombre, val = limpio(r[1]), pct_celda(r[2])
            if nombre and val is not None:
                adscritas.append({"nombre": nombre, "pct": val, "monto": num(r[3])})
        if adscritas:
            indicadores["adscritas"] = adscritas

    # Wasiymi (hojas INDICADORES WAYSIMI y POR REGION WAYSIMI — así escritas en el Excel)
    if "INDICADORES WAYSIMI" in wb.sheetnames:
        wasiymi = {}
        for r in wb["INDICADORES WAYSIMI"].iter_rows(values_only=True):
            etiqueta = str(r[1] or "").strip().upper()
            if etiqueta.startswith("META"):
                wasiymi["meta"] = num(r[2])
            elif etiqueta.startswith("EJECUCI"):
                wasiymi["ejecutadas"] = num(r[2])
        if wasiymi:
            indicadores["wasiymi"] = wasiymi
    if "POR REGION WAYSIMI" in wb.sheetnames and "wasiymi" in indicadores:
        reg_was = []
        for r in wb["POR REGION WAYSIMI"].iter_rows(values_only=True):
            nombre, val = limpio(r[1]), pct_celda(r[2])
            if nombre and val is not None:
                reg_was.append({"nombre": nombre, "pct": val})
        indicadores["wasiymi"]["regiones"] = reg_was

    # Regiones agrupadas por macro región (hoja POR REGION TITULOS)
    if "POR REGION TITULOS" in wb.sheetnames and "titulos" in indicadores:
        regiones, macro_actual = [], None
        for r in wb["POR REGION TITULOS"].iter_rows(values_only=True):
            nombre, val = limpio(r[1]), pct_celda(r[2])
            if not nombre:
                continue
            # cabecera de grupo (NORTE, CENTRO…): trae ejecución/meta en las
            # columnas 3-4 (los mismos totales de la hoja MACROREGION) o no trae %
            if num_celda(r[3]) is not None or val is None:
                macro_actual = nombre
            else:
                regiones.append({"macro": macro_actual, "nombre": nombre, "pct": val})
        indicadores["titulos"]["regiones"] = regiones

    with open(SALIDA, "w", encoding="utf-8") as f:
        f.write("window.PROYECTOS=")
        json.dump(proyectos, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";")
        f.write("window.INDICADORES=")
        json.dump(indicadores, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";")
    print("Indicadores:", indicadores)

    n = {}
    for p in proyectos:
        for k in p.get("carteras", {}):
            n[k] = n.get(k, 0) + 1
    print("Proyectos: %d | carteras: %s" % (len(proyectos), n))


if __name__ == "__main__":
    main()
